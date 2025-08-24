# analysis/html_route_processor.py (обновленный)
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import logging
import pandas as pd
import tempfile
import shutil
from typing import List, Dict, Tuple, Optional, Any
from collections import defaultdict
from bs4 import BeautifulSoup

# Настройка логирования
logger = logging.getLogger(__name__)

class HTMLRouteProcessor:
    """Процессор для обработки HTML файлов маршрутов с интеграцией route_processor.py"""
    
    def __init__(self):
        self.processed_routes = []
        self.processing_stats = {
            'total_files': 0,
            'total_routes_found': 0,
            'unique_routes': 0,
            'duplicates_total': 0,
            'routes_with_equal_rashod': 0,
            'routes_processed': 0,
            'routes_skipped': 0,
            'output_rows': 0,
            'duplicate_details': {}
        }
        self.routes_df = None
    
    # ================== УТИЛИТЫ ==================
    
    def clean_text(self, text: str) -> str:
        """Очищает текст от лишних пробелов и символов"""
        if not text:
            return ""
        text = text.replace('\xa0', ' ').replace('&nbsp;', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip()
    
    def try_convert_to_number(self, value: Any, force_int: bool = False) -> Optional[float]:
        """Преобразует значение в число, делая все числа положительными"""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        
        s = str(value).strip()
        s = s.replace(' ', '').replace('\xa0', '').replace('\u00a0', '')
        if s.endswith('.'):
            s = s[:-1]
        s = s.replace(',', '.')
        
        if s == '' or s == 'nan' or s.lower() == 'none':
            return None
        
        try:
            num = float(s)
            num = abs(num)  # Делаем число положительным
            if force_int or num == int(num):
                return int(num)
            return num
        except:
            return None
    
    def safe_subtract(self, *values) -> Optional[float]:
        """Безопасное вычитание с проверкой на None/NaN, возвращает абсолютное значение"""
        valid_values = [v for v in values if v is not None and not (isinstance(v, float) and pd.isna(v))]
        
        if not valid_values:
            return None
        
        result = valid_values[0]
        for v in valid_values[1:]:
            result = result - v
        
        return abs(result)
    
    def safe_divide(self, numerator: Any, denominator: Any) -> Optional[float]:
        """Безопасное деление с проверкой на None/NaN и деление на ноль"""
        if numerator is None or denominator is None:
            return None
        if isinstance(numerator, float) and pd.isna(numerator):
            return None
        if isinstance(denominator, float) and pd.isna(denominator):
            return None
        if denominator == 0:
            return None
        
        return abs(numerator / denominator)
    
    # ================== ПАРСИНГ HTML ==================
    
    def extract_norm_url_from_href(self, href: str) -> Optional[str]:
        """Извлекает номер нормы из URL гиперссылки"""
        if not href:
            return None
        
        try:
            from urllib.parse import urlparse, parse_qs
            parsed = urlparse(href)
            params = parse_qs(parsed.query)
            if 'id_ntp_tax' in params:
                return params['id_ntp_tax'][0]
        except Exception as e:
            logger.debug(f"Ошибка парсинга URL {href}: {e}")
        
        # Альтернативный метод через регулярное выражение
        match = re.search(r'id_ntp_tax=(\d+)', href)
        if match:
            return match.group(1)
        
        return None
    
    def extract_route_header_from_html(self, html_line: str) -> Optional[Dict]:
        """Извлекает метаданные маршрута из HTML строки"""
        soup = BeautifulSoup(html_line, 'html.parser')
        
        header = soup.find('th', class_='thl_common')
        if not header:
            return None
        
        header_text = header.get_text()
        metadata = {}
        
        # Номер маршрута
        route_spans = header.find_all('font', class_='filter_value')
        if route_spans and len(route_spans) > 0:
            metadata['number'] = self.clean_text(route_spans[0].get_text())
        else:
            patterns = [r'Маршрут\s*№[:\s]*(\d+)', r'Route\s*№[:\s]*(\d+)']
            for pattern in patterns:
                match = re.search(pattern, header_text)
                if match:
                    metadata['number'] = match.group(1)
                    break
        
        # Дата
        match = re.search(r'(\d{2}\.\d{2}\.\d{4})', header_text)
        metadata['date'] = match.group(1) if match else None
        
        # Депо
        depot_patterns = [r'Депо:\s*([^И]+)', r'Depot:\s*([^A-Z]+)']
        for pattern in depot_patterns:
            match = re.search(pattern, header_text)
            if match:
                metadata['depot'] = match.group(1).strip()
                break
        
        # Идентификатор
        id_patterns = [r'Идентификатор:\s*(\d+)', r'Identifier:\s*(\d+)']
        for pattern in id_patterns:
            match = re.search(pattern, header_text)
            if match:
                metadata['identifier'] = match.group(1)
                break
        
        # Извлекаем дополнительные данные из ТУ3
        trip_date, driver_tab = None, None
        try:
            loco_series, loco_number, trip_date, driver_tab = self.extract_loco_data_from_html(html_line)
            metadata['trip_date'] = trip_date
            metadata['driver_tab'] = driver_tab
        except:
            logger.debug("Не удалось извлечь дату поездки и табельный из ТУ3")
        
        return metadata
    
    def extract_loco_data_from_html(self, html_content: str) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
        """Извлекает серию и номер локомотива, дату поездки и табельный номер машиниста из HTML"""
        soup = BeautifulSoup(html_content, 'html.parser')
        
        tu3_fonts = soup.find_all('font', class_=['itog2', 'itog3'])
        
        for font in tu3_fonts:
            font_text = font.get_text().strip()
            if font_text in ['ТУ3', 'TU3'] or 'ТУ3' in font_text or 'TU3' in font_text:
                current = font
                data_elements = []
                
                while current:
                    current = current.find_next_sibling()
                    if current and current.name == 'font':
                        text = current.get_text().strip().replace('\xa0', ' ').replace('&nbsp;', ' ')
                        if text:
                            data_elements.append(text)
                    elif current and current.name == 'br':
                        break
                    elif not current:
                        break
                
                logger.debug(f"ТУ3 данные: {data_elements}")
                
                if len(data_elements) >= 7:
                    # Структура: [номер_маршрута, депо, номер_лок, серия, ..., дата_поездки, табельный, ...]
                    series_raw = data_elements[2]  # номер локомотиваseries_ra1w 
                    loco_number_raw = data_elements[3]       # серия
                    trip_date_raw = data_elements[4]    # дата поездки
                    driver_tab_raw = data_elements[5]   # табельный машиниста
                    
                    # Обрабатываем серию
                    if ',' in loco_number_raw:
                        series_part = loco_number_raw.split(',')[0]
                    else:
                        series_part = loco_number_raw
                    
                    series_clean = re.sub(r'[^\d]', '', series_part)
                    
                    # Убираем последнюю цифру если длина больше 1
                    if len(series_clean) > 1:
                        processed_series = series_clean[:-1]
                    else:
                        processed_series = series_clean
                    
                    # Меняем местами серию и номер как просил пользователь
                    final_loco_series = series_raw    # 240 -> серия
                    final_loco_number = processed_series   # 615 -> номер
                    
                    # Обрабатываем дату поездки - убираем все нецифровые символы
                    trip_date_clean = re.sub(r'[^\d]', '', trip_date_raw)
                    
                    # Обрабатываем табельный номер - убираем все нецифровые символы
                    driver_tab_clean = re.sub(r'[^\d]', '', driver_tab_raw)
                    
                    logger.debug(f"Результат: серия={final_loco_series}, номер={final_loco_number}, дата_поездки={trip_date_clean}, табельный={driver_tab_clean}")
                    return final_loco_series, final_loco_number, trip_date_clean, driver_tab_clean
        
        logger.debug("ТУ3 не найдено")
        return None, None, None, None
    
    def extract_yu7_data(self, html_content: str) -> List[Tuple[int, int, int]]:
        """Извлекает данные Ю7 из HTML"""
        soup = BeautifulSoup(html_content, 'html.parser')
        yu7_data = []
        
        all_fonts = soup.find_all('font')
        
        for font in all_fonts:
            font_text = font.get_text().strip()
            if any(pattern in font_text for pattern in ['Ю7', 'YU7', 'Yu7']):
                current = font
                data_elements = []
                
                while current:
                    current = current.find_next_sibling()
                    if current and current.name == 'font':
                        text = current.get_text().strip().replace('\xa0', ' ').replace('&nbsp;', ' ')
                        if text:
                            data_elements.append(text)
                    elif current and current.name == 'br':
                        break
                    elif not current:
                        break
                
                logger.debug(f"Ю7 данные найдены: {data_elements[:15]}...")
                
                # Структура данных Ю7: [строка, станция, время_приб, время_отпр, маневры, задержки, поезд, НЕТТО, БРУТТО, вагоны, ОСИ]
                if len(data_elements) >= 9:
                    try:
                        netto_text = data_elements[7].strip()
                        brutto_text = data_elements[8].strip()
                        
                        netto = int(netto_text)
                        brutto = int(brutto_text)
                        
                        # Ищем ОСИ
                        osi = None
                        for i in range(9, len(data_elements)):
                            element = data_elements[i]
                            if any(osi_pattern in element for osi_pattern in ['ОСИ', 'OSI']):
                                osi_match = re.search(r'(?:ОСИ|OSI)(\d+)', element)
                                if osi_match:
                                    osi = int(osi_match.group(1))
                                    break
                        
                        if osi is not None and brutto > 0:
                            yu7_data.append((netto, brutto, osi))
                            logger.debug(f"Ю7 успешно обработано: НЕТТО={netto}, БРУТТО={brutto}, ОСИ={osi}")
                    
                    except (ValueError, IndexError) as e:
                        logger.debug(f"Ошибка парсинга Ю7: {e}")
                        continue
        
        logger.debug(f"Всего найдено строк Ю7: {len(yu7_data)}")
        return yu7_data
    
    def find_matching_yu7(self, yu7_data: List[Tuple[int, int, int]], target_brutto: int, 
                         allow_double: bool = True, tolerance_percent: float = 5.0) -> Tuple[Optional[Tuple[int, int, int]], bool, bool]:
        """Находит первую строку Ю7 с подходящим БРУТТО"""
        logger.debug(f"Поиск Ю7 для target_brutto={target_brutto}")
        
        # ЭТАП 1: Точное совпадение
        for netto, brutto, osi in yu7_data:
            if brutto == target_brutto:
                logger.info(f"✓ Найдено точное совпадение: НЕТТО={netto}, БРУТТО={brutto}, ОСИ={osi}")
                return (netto, brutto, osi), False, False
        
        # ЭТАП 2: Совпадение с допустимой погрешностью до 5%
        best_match = None
        min_percentage_diff = float('inf')
        
        for netto, brutto, osi in yu7_data:
            diff = abs(brutto - target_brutto)
            percentage_diff = (diff / target_brutto) * 100
            
            if percentage_diff <= tolerance_percent and percentage_diff < min_percentage_diff:
                min_percentage_diff = percentage_diff
                best_match = (netto, brutto, osi)
        
        if best_match:
            netto, brutto, osi = best_match
            logger.info(f"✓ Найдено приближенное совпадение: target={target_brutto}, найденное БРУТТО={brutto}, отклонение={min_percentage_diff:.2f}%")
            return best_match, False, True
        
        # ЭТАП 3: Двойная тяга
        if allow_double:
            double_target = target_brutto * 2
            logger.debug(f"Поиск двойной тяги для target_brutto={double_target}")
            
            # Точное совпадение для двойной тяги
            for netto, brutto, osi in yu7_data:
                if brutto == double_target:
                    logger.info(f"✓ Найдена точная двойная тяга: НЕТТО={netto}, БРУТТО={brutto}, ОСИ={osi}")
                    return (netto, brutto, osi), True, False
            
            # Приближенное совпадение для двойной тяги
            best_double_match = None
            min_double_percentage_diff = float('inf')
            
            for netto, brutto, osi in yu7_data:
                diff = abs(brutto - double_target)
                percentage_diff = (diff / double_target) * 100
                
                if percentage_diff <= tolerance_percent and percentage_diff < min_double_percentage_diff:
                    min_double_percentage_diff = percentage_diff
                    best_double_match = (netto, brutto, osi)
            
            if best_double_match:
                netto, brutto, osi = best_double_match
                logger.info(f"✓ Найдена приближенная двойная тяга: отклонение={min_double_percentage_diff:.2f}%")
                return best_double_match, True, True
        
        logger.warning(f"✗ Совпадение не найдено для target_brutto={target_brutto}")
        return None, False, False
    
    # ================== ОБРАБОТКА УЧАСТКОВ ==================
    
    def can_merge_sections(self, sections: List[Dict]) -> bool:
        """Проверяет, можно ли объединить участки с одинаковым названием"""
        norm_numbers = []
        for section in sections:
            norm_num = section.get('norm_number')
            if norm_num is None:
                ud_norma_url = section.get('ud_norma_url')
                if ud_norma_url:
                    norm_num = self.extract_norm_url_from_href(ud_norma_url)
            
            if norm_num is not None:
                norm_numbers.append(norm_num)
        
        unique_norms = set(norm_numbers)
        can_merge = len(unique_norms) <= 1
        
        logger.debug(f"Проверка возможности объединения: номера норм {norm_numbers}, можно объединить: {can_merge}")
        return can_merge
    
    def get_merged_norm_number(self, sections: List[Dict]) -> Optional[str]:
        """Возвращает номер нормы для объединенного участка"""
        for section in sections:
            norm_num = section.get('norm_number')
            if norm_num is not None:
                return norm_num
        
        for section in sections:
            ud_norma_url = section.get('ud_norma_url')
            if ud_norma_url:
                norm_num = self.extract_norm_url_from_href(ud_norma_url)
                if norm_num:
                    return norm_num
        
        return None
    
    def merge_sections(self, sections: List[Dict], yu7_data: List[Tuple[int, int, int]]) -> Optional[Dict]:
        """Объединяет участки с одинаковым названием"""
        if not sections:
            return None
        
        section_name = sections[0].get('name')
        logger.debug(f"Попытка объединения участка '{section_name}': {len(sections)} участков")
        
        if not self.can_merge_sections(sections):
            logger.debug(f"Нельзя объединить участки '{section_name}': разные номера норм")
            return None
        
        # Берем данные первого участка как основу
        merged = sections[0].copy()
        
        # Поля для суммирования
        sum_fields = [
            'tkm_brutto', 'km', 'rashod_fact', 'rashod_norm',
            'prostoy_vsego', 'prostoy_norma', 'manevry_vsego', 'manevry_norma',
            'troganie_vsego', 'troganie_norma', 'nagon_vsego', 'nagon_norma',
            'ogranich_vsego', 'ogranich_norma', 'peresyl_vsego', 'peresyl_norma'
        ]
        
        # Суммируем указанные поля
        for field in sum_fields:
            total = 0
            has_values = False
            for section in sections:
                value = section.get(field)
                if value is not None and value != '':
                    try:
                        float_val = float(value)
                        total += float_val
                        has_values = True
                    except (ValueError, TypeError):
                        pass
            
            if has_values:
                merged[field] = total
            else:
                merged[field] = None
        
        # Устанавливаем номер нормы
        merged['norm_number'] = self.get_merged_norm_number(sections)
        
        # Пытаемся найти соответствующие НЕТТО, БРУТТО, ОСИ
        tkm_brutto = merged.get('tkm_brutto')
        km = merged.get('km')
        
        if tkm_brutto and km and tkm_brutto > 0 and km > 0:
            target_brutto_float = tkm_brutto / km
            target_brutto = round(target_brutto_float)
            logger.debug(f"Целевое БРУТТО для '{section_name}': {target_brutto}")
            
            matched, is_double, is_approximate = self.find_matching_yu7(yu7_data, target_brutto, allow_double=True)
            
            if matched:
                netto, brutto, osi = matched
                merged['netto'] = netto
                merged['brutto'] = brutto
                merged['osi'] = osi
                merged['use_red_color'] = is_approximate
                merged['double_traction'] = "Да" if is_double else None
                merged['is_merged'] = True
                
                logger.info(f"✓ Объединение участка '{section_name}' успешно")
                return merged
            else:
                # Если соответствующие Ю7 данные не найдены, устанавливаем красные "-"
                merged['netto'] = "-"
                merged['brutto'] = "-"
                merged['osi'] = "-"
                merged['use_red_color'] = True
                merged['double_traction'] = None
                merged['is_merged'] = True
                
                logger.warning(f"✗ Для объединенного участка '{section_name}' не найдены Ю7 данные")
                return merged
        else:
            logger.warning(f"✗ Недостаточно данных для объединения участка '{section_name}'")
            return None
    
    def merge_identical_sections(self, norm_sections: List[Dict], station_sections: Dict[str, Dict], 
                               yu7_data: List[Tuple[int, int, int]]) -> List[Dict]:
        """Объединяет одинаковые участки в рамках одного маршрута"""
        logger.info("Начинаем объединение одинаковых участков")
        
        # Группируем участки по названию
        sections_by_name = defaultdict(list)
        for section in norm_sections:
            section_name = section.get('name')
            if section_name:
                sections_by_name[section_name].append(section)
        
        merged_sections = []
        
        for section_name, sections in sections_by_name.items():
            if len(sections) == 1:
                merged_sections.append(sections[0])
            else:
                logger.info(f"Участок '{section_name}': найдено {len(sections)} одинаковых")
                
                # Добавляем данные станций к каждому участку перед объединением
                for section in sections:
                    station_data = station_sections.get(section_name, {})
                    section.update({
                        'prostoy_vsego': station_data.get('prostoy_vsego'),
                        'prostoy_norma': station_data.get('prostoy_norma'),
                        'manevry_vsego': station_data.get('manevry_vsego'),
                        'manevry_norma': station_data.get('manevry_norma'),
                        'troganie_vsego': station_data.get('troganie_vsego'),
                        'troganie_norma': station_data.get('troganie_norma'),
                        'nagon_vsego': station_data.get('nagon_vsego'),
                        'nagon_norma': station_data.get('nagon_norma'),
                        'ogranich_vsego': station_data.get('ogranich_vsego'),
                        'ogranich_norma': station_data.get('ogranich_norma'),
                        'peresyl_vsego': station_data.get('peresyl_vsego'),
                        'peresyl_norma': station_data.get('peresyl_norma')
                    })
                
                merged_section = self.merge_sections(sections, yu7_data)
                
                if merged_section:
                    merged_sections.append(merged_section)
                else:
                    merged_sections.extend(sections)
        
        logger.info(f"Результат объединения: было {len(norm_sections)} участков, стало {len(merged_sections)}")
        return merged_sections
    
    # ================== ПАРСИНГ ТАБЛИЦ ==================
    
    def parse_norm_table(self, soup: BeautifulSoup) -> List[Dict]:
        """Парсит таблицу с нормами"""
        norm_sections = []
        
        tables = soup.find_all('table')
        
        for table in tables:
            headers = table.find_all('th')
            header_texts = [self.clean_text(h.get_text()) for h in headers]
            
            # Проверяем, что это нужная таблица
            if any('Нормируемый участок' in h for h in header_texts):
                # Определяем индексы колонок
                col_indices = {}
                for idx, header in enumerate(header_texts):
                    header_lower = header.lower()
                    if 'участок' in header_lower and 'станция' in header_lower:
                        col_indices['name'] = idx
                    elif 'ткм брутто' in header_lower:
                        col_indices['tkm_brutto'] = idx
                    elif header_lower == 'км' or header_lower == 'км.':
                        col_indices['km'] = idx
                    elif header_lower == 'пр.' or header_lower == 'пр':
                        col_indices['pr'] = idx
                    elif 'расход фактический' in header_lower:
                        col_indices['rashod_fact'] = idx
                    elif 'расход по норме' in header_lower:
                        col_indices['rashod_norm'] = idx
                    elif 'уд. норма' in header_lower or 'норма на 1 час ман. раб.' in header_lower:
                        col_indices['ud_norma'] = idx
                    elif 'норма на работу' in header_lower:
                        col_indices['norma_rabotu'] = idx
                    elif 'норма на одиночное' in header_lower:
                        col_indices['norma_odinochnoe'] = idx
                
                # Парсим строки данных
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    if not cells:
                        continue
                    
                    # Проверяем, что это не итоговая строка
                    first_cell_text = self.clean_text(cells[0].get_text())
                    if 'итого' in first_cell_text.lower():
                        continue
                    
                    section = {}
                    
                    # Извлекаем данные по индексам
                    for field, idx in col_indices.items():
                        if idx < len(cells):
                            cell = cells[idx]
                            value = self.clean_text(cell.get_text())
                            
                            # Для удельной нормы также извлекаем URL и номер нормы
                            if field == 'ud_norma':
                                link = cell.find('a')
                                if link and link.get('href'):
                                    section['ud_norma_url'] = link.get('href')
                                    norm_number = self.extract_norm_url_from_href(link.get('href'))
                                    if norm_number:
                                        section['norm_number'] = norm_number
                            
                            # Преобразуем числовые значения
                            if field != 'name':
                                value = self.try_convert_to_number(value)
                            
                            section[field] = value
                    
                    if section.get('name'):
                        norm_sections.append(section)
        
        return norm_sections
    
    def parse_station_table(self, soup: BeautifulSoup) -> Dict[str, Dict]:
        """Парсит таблицу со станциями и дополнительными данными"""
        station_sections = {}
        
        tables = soup.find_all('table')
        
        for table in tables:
            headers = table.find_all('th')
            header_texts = [self.clean_text(h.get_text()) for h in headers]
            
            # Проверяем, что это нужная таблица
            patterns = ['В том числе', 'In that number']
            if any(any(pattern in h for pattern in patterns) for h in header_texts):
                rows = table.find_all('tr')
                
                for row in rows:
                    cells = row.find_all('td')
                    if not cells:
                        continue
                    
                    # Первая ячейка - название участка
                    section_name = self.clean_text(cells[0].get_text())
                    if not section_name or any(pattern in section_name.lower() for pattern in ['итого', 'total']):
                        continue
                    
                    # Парсим остальные данные
                    data = {}
                    
                    if len(cells) > 1:
                        data['prostoy_vsego'] = self.try_convert_to_number(cells[1].get_text())
                    if len(cells) > 2:
                        data['prostoy_norma'] = self.try_convert_to_number(cells[2].get_text())
                    if len(cells) > 3:
                        data['manevry_vsego'] = self.try_convert_to_number(cells[3].get_text())
                    if len(cells) > 4:
                        data['manevry_norma'] = self.try_convert_to_number(cells[4].get_text())
                    if len(cells) > 5:
                        data['troganie_vsego'] = self.try_convert_to_number(cells[5].get_text())
                    if len(cells) > 6:
                        data['troganie_norma'] = self.try_convert_to_number(cells[6].get_text())
                    if len(cells) > 7:
                        data['nagon_vsego'] = self.try_convert_to_number(cells[7].get_text())
                    if len(cells) > 8:
                        data['nagon_norma'] = self.try_convert_to_number(cells[8].get_text())
                    if len(cells) > 9:
                        data['ogranich_vsego'] = self.try_convert_to_number(cells[9].get_text())
                    if len(cells) > 10:
                        data['ogranich_norma'] = self.try_convert_to_number(cells[10].get_text())
                    if len(cells) > 11:
                        data['peresyl_vsego'] = self.try_convert_to_number(cells[11].get_text())
                    if len(cells) > 12:
                        data['peresyl_norma'] = self.try_convert_to_number(cells[12].get_text())
                    
                    station_sections[section_name] = data
        
        return station_sections
    
    def calculate_fact_na_rabotu(self, section: Dict, station_data: Dict) -> Optional[float]:
        """Вычисляет Факт на работу"""
        rashod_fact = section.get('rashod_fact')
        if rashod_fact is None:
            return None
        
        # Вычитаем нормы
        prostoy_norma = station_data.get('prostoy_norma')
        troganie_norma = station_data.get('troganie_norma')
        nagon_norma = station_data.get('nagon_norma')
        ogranich_norma = station_data.get('ogranich_norma')
        peresyl_norma = station_data.get('peresyl_norma')
        
        return self.safe_subtract(
            rashod_fact,
            prostoy_norma,
            troganie_norma,
            nagon_norma,
            ogranich_norma,
            peresyl_norma
        )
    
    def calculate_fact_ud(self, fact_na_rabotu: Optional[float], 
                         tkm_brutto: Optional[float]) -> Optional[float]:
        """Вычисляет Факт уд"""
        if fact_na_rabotu is None or tkm_brutto is None:
            return None
        
        tkm_10000 = self.safe_divide(tkm_brutto, 10000)
        if tkm_10000 is None:
            return None
        
        return self.safe_divide(fact_na_rabotu, tkm_10000)
    
    # ================== ГЛАВНЫЕ ФУНКЦИИ ПАРСИНГА ==================
    
    def clean_html_file(self, input_file: str) -> str:
        """Очищает HTML файл от лишнего кода (адаптация из delet_code_marhrut.py)"""
        logger.info(f"Очистка HTML файла: {input_file}")
        
        # Проверяем существование файла
        if not os.path.exists(input_file):
            logger.error(f"Файл {input_file} не найден!")
            return None
        
        try:
            # Читаем файл с кодировкой cp1251
            with open(input_file, 'r', encoding='cp1251') as f:
                content = f.read()
            encoding_used = 'cp1251'
            logger.debug(f"Файл прочитан с кодировкой cp1251, размер: {len(content)} байт")
        except UnicodeDecodeError:
            # Пробуем другие кодировки
            try:
                with open(input_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                encoding_used = 'utf-8'
                logger.debug(f"Файл прочитан с кодировкой utf-8")
            except UnicodeDecodeError:
                with open(input_file, 'r', encoding='latin-1') as f:
                    content = f.read()
                encoding_used = 'latin-1'
                logger.debug(f"Файл прочитан с кодировкой latin-1")
        
        # Ищем начальный маркер
        start_marker = '<table align=center width="100%">'
        start_pos = content.find(start_marker)
        
        if start_pos == -1:
            logger.error(f"Начальный маркер не найден в файле {input_file}")
            return None
        
        # Ищем конечный маркер
        form_pattern = r'</table>\s*</td>\s*</tr></table><form id=print_form>.*?(?=\n|$)'
        form_match = re.search(form_pattern, content[start_pos:], re.DOTALL)
        
        if not form_match:
            logger.error(f"Конечный маркер не найден в файле {input_file}")
            return None
        
        # Позиция конца найденной строки с формой
        form_end_pos = start_pos + form_match.end()
        
        # Ищем следующую строку '</td></tr>' после формы
        remaining_content = content[form_end_pos:]
        lines = remaining_content.split('\n')
        
        end_pos = form_end_pos
        for i, line in enumerate(lines):
            if '</td></tr>' in line.strip():
                end_pos = form_end_pos + len('\n'.join(lines[:i+1]))
                break
        
        # Извлекаем нужную часть
        extracted_data = content[start_pos:end_pos + 1]
        
        # Разбиваем маршруты по отдельным строкам
        extracted_data = self._split_routes_to_lines(extracted_data)
        
        # Удаляем маршруты с " ВЧТ "
        extracted_data = self._remove_vcht_routes(extracted_data)
        
        # Очищаем HTML код от лишних элементов
        extracted_data = self._clean_html_content(extracted_data)
        
        # Создаем временный файл
        temp_file = tempfile.mktemp(suffix='.html')
        with open(temp_file, 'w', encoding=encoding_used) as f:
            f.write(extracted_data)
        
        logger.info(f"HTML файл очищен и сохранен во временный файл: {temp_file}")
        return temp_file
    
    def _split_routes_to_lines(self, content: str) -> str:
        """Разбивает содержимое HTML на отдельные строки для каждого маршрута"""
        logger.debug("Разбиваем маршруты по отдельным строкам...")
        
        # Паттерн для поиска полного маршрута от начала до конца
        route_pattern = r'(<table[^>]*><tr><th class=thl_common><font class=filter_key>\s*Маршрут\s*№:.*?<br><br><br>)'
        
        # Находим все маршруты
        routes = re.findall(route_pattern, content, flags=re.DOTALL)
        
        if not routes:
            logger.warning("Маршруты не найдены для разделения")
            return content
        
        logger.debug(f"Найдено маршрутов для разделения: {len(routes)}")
        
        # Находим код до первого маршрута
        first_route_start = content.find(routes[0])
        before_routes = content[:first_route_start]
        
        # Находим код после последнего маршрута  
        last_route_end = content.rfind(routes[-1]) + len(routes[-1])
        after_routes = content[last_route_end:]
        
        # Собираем результат
        result_lines = []
        
        if before_routes.strip():
            result_lines.append(before_routes.rstrip())
        
        result_lines.append("<!-- НАЧАЛО_ПЕРВОГО_МАРШРУТА -->")
        
        for route in routes:
            result_lines.append(route)
        
        result_lines.append("<!-- КОНЕЦ_ПОСЛЕДНЕГО_МАРШРУТА -->")
        
        if after_routes.strip():
            result_lines.append(after_routes.lstrip())
        
        logger.debug(f"Маршруты разделены на {len(routes)} отдельных строк")
        return '\n'.join(result_lines)
    
    def _remove_vcht_routes(self, content: str) -> str:
        """Удаляет строки с маршрутами, содержащими ' ВЧТ '"""
        logger.debug("Удаляем маршруты с ' ВЧТ '...")
        
        lines = content.split('\n')
        filtered_lines = []
        removed_routes = 0
        
        for line in lines:
            if '<td class = itog2>" ВЧТ "</td>' in line:
                removed_routes += 1
                logger.debug(f"Удален маршрут с ВЧТ (строка {len(filtered_lines) + 1})")
                continue
            filtered_lines.append(line)
        
        if removed_routes > 0:
            logger.info(f"Удалено {removed_routes} маршрутов с ' ВЧТ '")
        
        return '\n'.join(filtered_lines)
    
    def _clean_html_content(self, content: str) -> str:
        """Очищает HTML контент от лишних элементов"""
        logger.debug("Очищаем HTML код от лишних элементов...")
        
        original_size = len(content)
        
        # Удаляем лишние элементы (как в оригинальном коде)
        date_pattern = r'<font class = rcp12 ><center>Дата получения:.*?</font>\s*<br>'
        content = re.sub(date_pattern, '', content, flags=re.DOTALL)
        
        route_num_pattern = r'<font class = rcp12 ><center>Номер маршрута:.*?</font><br>'
        content = re.sub(route_num_pattern, '', content, flags=re.DOTALL)
        
        numline_pattern = r'<tr class=tr_numline>.*?</tr>'
        content = re.sub(numline_pattern, '', content, flags=re.DOTALL)
        
        # Удаляем атрибуты выравнивания и другие лишние элементы
        content = re.sub(r'\s+ALIGN=center', '', content, flags=re.IGNORECASE)
        content = re.sub(r'\s+align=left', '', content, flags=re.IGNORECASE)
        content = re.sub(r'\s+align=right', '', content, flags=re.IGNORECASE)
        content = re.sub(r'<center>', '', content)
        content = re.sub(r'</center>', '', content)
        content = re.sub(r'<pre>', '', content)
        content = re.sub(r'</pre>', '', content)
        content = re.sub(r'>[ \t]+<', '><', content)
        
        cleaned_size = len(content)
        removed_bytes = original_size - cleaned_size
        
        logger.debug(f"Удалено {removed_bytes:,} байт лишнего кода ({removed_bytes/original_size*100:.1f}%)")
        
        return content
    
    def extract_routes_from_html(self, html_content: str) -> List[Tuple[str, Dict]]:
        """Извлекает маршруты из HTML контента"""
        logger.info("Начинаем извлечение маршрутов из HTML")
        
        # Ищем маркеры начала и конца маршрутов
        start_marker = "<!-- НАЧАЛО_ПЕРВОГО_МАРШРУТА -->"
        end_marker = "<!-- КОНЕЦ_ПОСЛЕДНЕГО_МАРШРУТА -->"
        
        start_pos = html_content.find(start_marker)
        end_pos = html_content.find(end_marker)
        
        if start_pos == -1 or end_pos == -1:
            logger.warning("Маркеры маршрутов не найдены, обрабатываем весь контент")
            routes_section = html_content
        else:
            routes_section = html_content[start_pos + len(start_marker):end_pos]
        
        # Разбиваем на строки
        lines = routes_section.strip().split('\n')
        routes = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Более гибкий поиск маршрутов - любая ширина table
            if re.search(r'<table width=\d+%', line) and ('Маршрут №' in line or 'Маршрут' in line):
                # Извлекаем метаданные маршрута
                metadata = self.extract_route_header_from_html(line)
                if metadata:
                    routes.append((line, metadata))
                    logger.debug(f"Найден маршрут: №{metadata['number']}")
        
        logger.info(f"Найдено маршрутов: {len(routes)}")
        return routes
    
    def check_rashod_equal_html(self, route_html: str) -> bool:
        """Проверяет равны ли Расход по норме и Расход фактический в маршруте"""
        soup = BeautifulSoup(route_html, 'html.parser')
        
        norm_tables = soup.find_all('table', width='90%')
        
        for table in norm_tables:
            headers = table.find_all('th')
            header_texts = [self.clean_text(h.get_text()) for h in headers]
            
            if any('Нормируемый участок' in h for h in header_texts):
                rashod_fact_idx = None
                rashod_norm_idx = None
                
                for idx, header in enumerate(header_texts):
                    if 'Расход фактический' in header:
                        rashod_fact_idx = idx
                    elif 'Расход по норме' in header:
                        rashod_norm_idx = idx
                
                if rashod_fact_idx is not None and rashod_norm_idx is not None:
                    rows = table.find_all('tr')
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) > max(rashod_fact_idx, rashod_norm_idx):
                            fact_text = self.clean_text(cells[rashod_fact_idx].get_text())
                            norm_text = self.clean_text(cells[rashod_norm_idx].get_text())
                            
                            fact_val = self.try_convert_to_number(fact_text)
                            norm_val = self.try_convert_to_number(norm_text)
                            
                            if fact_val is not None and norm_val is not None:
                                if abs(fact_val - norm_val) < 0.01:
                                    return True
        
        return False
    
    def select_best_route(self, routes: List[Tuple[str, Dict]]) -> Optional[Tuple[str, Dict]]:
        """Выбирает лучший маршрут из списка дубликатов"""
        if not routes:
            return None
        
        logger.debug(f"Выбор лучшего маршрута из {len(routes)} вариантов")
        
        # Фильтруем маршруты с равными расходами
        valid_routes = []
        equal_routes = []
        
        for route_html, metadata in routes:
            is_equal = self.check_rashod_equal_html(route_html)
            if is_equal:
                equal_routes.append((route_html, metadata))
            else:
                valid_routes.append((route_html, metadata))
        
        # Если есть валидные маршруты, выбираем с максимальным ID
        if valid_routes:
            best_route = max(valid_routes, key=lambda x: int(x[1]['identifier']) if x[1]['identifier'] else 0)
            return best_route
        
        # Если все маршруты с равными расходами, берем последний
        if equal_routes:
            best_route = max(equal_routes, key=lambda x: int(x[1]['identifier']) if x[1]['identifier'] else 0)
            return best_route
        
        return None
    
    def parse_html_route(self, route_html: str, metadata: Dict, 
                        has_equal_duplicates: bool = False, 
                        rashod_equal: bool = False) -> List[Dict]:
        """Главная функция парсинга HTML маршрута"""
        soup = BeautifulSoup(route_html, 'html.parser')
        
        # Базовые данные маршрута
        route_number = metadata.get('number')
        route_date = metadata.get('date')
        depot = metadata.get('depot', '')
        identifier = metadata.get('identifier')
        
        # Извлекаем серию и номер локомотива, дату поездки и табельный
        loco_series, loco_number, trip_date, driver_tab = self.extract_loco_data_from_html(route_html)
        
        # Извлекаем данные Ю7 (НЕТТО, БРУТТО, ОСИ)
        yu7_data = self.extract_yu7_data(route_html)
        default_netto = yu7_data[0][0] if yu7_data else None
        default_brutto = yu7_data[0][1] if yu7_data else None
        default_osi = yu7_data[0][2] if yu7_data else None
        
        # Парсим таблицы с нормами
        norm_sections = self.parse_norm_table(soup)
        station_sections = self.parse_station_table(soup)
        
        # Объединяем одинаковые участки
        merged_sections = self.merge_identical_sections(norm_sections, station_sections, yu7_data)
        
        # Формируем результат
        output_rows = []
        
        for section in merged_sections:
            section_name = section.get('name')
            if not section_name:
                continue
            
            # Получаем данные станции (если участок не был объединен)
            if not section.get('is_merged'):
                station_data = station_sections.get(section_name, {})
            else:
                station_data = {}  # Для объединенных участков данные уже включены
            
            # Извлекаем номер нормы
            norm_number = section.get('norm_number')
            if not norm_number:
                ud_norma_url = section.get('ud_norma_url')
                if ud_norma_url:
                    norm_number = self.extract_norm_url_from_href(ud_norma_url)
            
            # Определяем НЕТТО, БРУТТО, ОСИ для участка
            netto = section.get('netto')
            brutto = section.get('brutto')
            osi = section.get('osi')
            use_red_color = section.get('use_red_color', False)
            double_traction = section.get('double_traction')
            
            # Если данные не определены, используем стандартную логику
            if netto is None or brutto is None or osi is None:
                tkm_brutto = section.get('tkm_brutto')
                km = section.get('km')
                
                if tkm_brutto and km and tkm_brutto > 0 and km > 0:
                    target_brutto_float = tkm_brutto / km
                    target_brutto = round(target_brutto_float)
                    
                    # Ищем подходящую строку Ю7
                    matched, is_double, is_approximate = self.find_matching_yu7(yu7_data, target_brutto, allow_double=True)
                    
                    if matched:
                        netto, brutto, osi = matched
                        use_red_color = is_approximate
                        if is_double:
                            double_traction = "Да"
                    else:
                        netto = "-"
                        brutto = "-"
                        osi = "-"
                        use_red_color = True
                else:
                    netto = default_netto
                    brutto = default_brutto
                    osi = default_osi
            
            # Вычисляем нажатие на ось
            axle_load = None
            if (brutto and osi and brutto != "-" and osi != "-" and 
                not use_red_color and isinstance(brutto, (int, float)) and isinstance(osi, (int, float))):
                axle_load = brutto / osi
            else:
                if brutto == "-" or osi == "-":
                    axle_load = "-"
            
            # Вычисляем Факт на работу и Факт уд
            fact_na_rabotu = self.calculate_fact_na_rabotu(section, station_data if not section.get('is_merged') else section)
            fact_ud = self.calculate_fact_ud(fact_na_rabotu, section.get('tkm_brutto'))
            
            # Формируем строку данных
            row = {
                'Номер маршрута': route_number,
                'Дата маршрута': route_date,
                'Дата поездки': trip_date,
                'Табельный машиниста': driver_tab,
                'Депо': depot,
                'Идентификатор': identifier,
                'Серия локомотива': loco_series,
                'Номер локомотива': loco_number,
                'НЕТТО': netto,
                'БРУТТО': brutto,
                'ОСИ': osi,
                'USE_RED_COLOR': use_red_color,
                'USE_RED_RASHOD': rashod_equal,
                'Наименование участка': section_name,
                'Номер нормы': norm_number,
                'Дв. тяга': double_traction,
                'Ткм брутто': section.get('tkm_brutto'),
                'Км': section.get('km'),
                'Пр.': section.get('pr'),
                'Расход фактический': section.get('rashod_fact'),
                'Расход по норме': section.get('rashod_norm'),
                'Уд. норма, норма на 1 час ман. раб.': section.get('ud_norma'),
                'Нажатие на ось': axle_load,
                'Норма на работу': section.get('norma_rabotu'),
                'Факт уд': fact_ud,
                'Факт на работу': fact_na_rabotu,
                'Норма на одиночное': section.get('norma_odinochnoe'),
                'Простой с бригадой, мин., всего': section.get('prostoy_vsego') if section.get('is_merged') else station_data.get('prostoy_vsego'),
                'Простой с бригадой, мин., норма': section.get('prostoy_norma') if section.get('is_merged') else station_data.get('prostoy_norma'),
                'Маневры, мин., всего': section.get('manevry_vsego') if section.get('is_merged') else station_data.get('manevry_vsego'),
                'Маневры, мин., норма': section.get('manevry_norma') if section.get('is_merged') else station_data.get('manevry_norma'),
                'Трогание с места, случ., всего': section.get('troganie_vsego') if section.get('is_merged') else station_data.get('troganie_vsego'),
                'Трогание с места, случ., норма': section.get('troganie_norma') if section.get('is_merged') else station_data.get('troganie_norma'),
                'Нагон опозданий, мин., всего': section.get('nagon_vsego') if section.get('is_merged') else station_data.get('nagon_vsego'),
                'Нагон опозданий, мин., норма': section.get('nagon_norma') if section.get('is_merged') else station_data.get('nagon_norma'),
                'Ограничения скорости, случ., всего': section.get('ogranich_vsego') if section.get('is_merged') else station_data.get('ogranich_vsego'),
                'Ограничения скорости, случ., норма': section.get('ogranich_norma') if section.get('is_merged') else station_data.get('ogranich_norma'),
                'На пересылаемые л-вы, всего': section.get('peresyl_vsego') if section.get('is_merged') else station_data.get('peresyl_vsego'),
                'На пересылаемые л-вы, норма': section.get('peresyl_norma') if section.get('is_merged') else station_data.get('peresyl_norma'),
                'Количество дубликатов маршрута': '',
                'Н=Ф': 'Да' if has_equal_duplicates else None
            }
            
            output_rows.append(row)
        
        return output_rows
    
    # ================== ОСНОВНЫЕ ФУНКЦИИ ПРОЦЕССОРА ==================
    
    def process_html_files(self, html_files: List[str]) -> pd.DataFrame:
        """Обрабатывает список HTML файлов маршрутов"""
        logger.info(f"Начинаем обработку {len(html_files)} HTML файлов")
        
        self.processing_stats['total_files'] = len(html_files)
        all_routes_data = []
        
        # Обрабатываем каждый файл
        for file_path in html_files:
            logger.info(f"Обработка файла: {os.path.basename(file_path)}")
            
            try:
                # Очищаем HTML файл
                cleaned_file = self.clean_html_file(file_path)
                if not cleaned_file:
                    logger.error(f"Не удалось очистить файл {file_path}")
                    continue
                
                # Обрабатываем очищенный файл
                file_routes = self._process_single_cleaned_file(cleaned_file)
                all_routes_data.extend(file_routes)
                
                # Удаляем временный файл
                try:
                    os.remove(cleaned_file)
                except:
                    pass
                
            except Exception as e:
                logger.error(f"Ошибка при обработке файла {file_path}: {e}")
                continue
        
        # Фильтруем и выбираем лучшие маршруты
        if all_routes_data:
            df = self._filter_and_select_best_routes(all_routes_data)
            logger.info(f"Обработка завершена. Получено {len(df)} итоговых записей")
            
            # Сохраняем DataFrame в класс
            self.routes_df = df
            return df
        else:
            logger.warning("Не получено ни одной записи из всех файлов")
            return pd.DataFrame()
    
    def _process_single_cleaned_file(self, cleaned_file: str) -> List[Dict]:
        """Обрабатывает один очищенный HTML файл"""
        logger.debug(f"Обработка очищенного файла: {cleaned_file}")
        
        try:
            with open(cleaned_file, 'r', encoding='cp1251') as f:
                html_content = f.read()
        except:
            with open(cleaned_file, 'r', encoding='utf-8') as f:
                html_content = f.read()
        
        # Обрабатываем маршруты
        df, stats = self._process_routes(html_content)
        
        # Обновляем общую статистику
        self.processing_stats['total_routes_found'] += stats['total_routes_found']
        self.processing_stats['unique_routes'] += stats['unique_routes']
        self.processing_stats['duplicates_total'] += stats['duplicates_total']
        self.processing_stats['routes_with_equal_rashod'] += stats['routes_with_equal_rashod']
        self.processing_stats['routes_processed'] += stats['routes_processed']
        self.processing_stats['routes_skipped'] += stats['routes_skipped']
        self.processing_stats['output_rows'] += stats['output_rows']
        
        if not df.empty:
            return df.to_dict('records')
        else:
            return []
    
    def _process_routes(self, html_content: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Обрабатывает все маршруты из HTML"""
        
        # Извлекаем маршруты
        routes = self.extract_routes_from_html(html_content)
        
        # Инициализируем статистику
        stats = {
            'total_routes_found': len(routes),
            'unique_routes': 0,
            'duplicates_total': 0,
            'routes_with_equal_rashod': 0,
            'routes_skipped': 0,
            'routes_processed': 0,
            'output_rows': 0,
            'duplicate_details': {}
        }
        
        if not routes:
            logger.error("Маршруты не найдены в HTML")
            return pd.DataFrame(), stats
        
        # Группируем маршруты по новому ключу (номер + дата поездки + табельный)
        route_groups = defaultdict(list)
        skipped_by_yu6 = 0
        
        for route_html, metadata in routes:
            # Проверяем на Ю6 фильтр
            if self.check_yu6_filter(route_html):
                skipped_by_yu6 += 1
                stats['routes_skipped'] += 1
                continue
            
            if metadata['number'] and metadata.get('trip_date') and metadata.get('driver_tab'):
                key = f"{metadata['number']}_{metadata.get('trip_date')}_{metadata.get('driver_tab')}"
                route_groups[key].append((route_html, metadata))
            else:
                stats['routes_skipped'] += 1
        
        if skipped_by_yu6 > 0:
            logger.info(f"Пропущено {skipped_by_yu6} маршрутов по фильтру Ю6")
        
        stats['unique_routes'] = len(route_groups)
        
        # Обрабатываем каждую группу
        all_rows = []
        duplicate_counts = {}
        
        for key, group in route_groups.items():
            logger.info(f"Обработка маршрута {key}, версий: {len(group)}")
            
            # Подсчитываем дубликаты
            if len(group) > 1:
                duplicate_counts[key] = len(group) - 1
                stats['duplicates_total'] += len(group) - 1
                stats['duplicate_details'][key] = {
                    'versions': len(group),
                    'duplicates': len(group) - 1,
                    'identifiers': [g[1]['identifier'] for g in group]
                }
            
            # Проверяем маршруты с равными расходами (старая логика, но теперь не используется для исключения)
            equal_count = sum(1 for r in group if self.check_rashod_equal_html(r[0]))
            if equal_count > 0:
                stats['routes_with_equal_rashod'] += equal_count
            
            # Выбираем лучший маршрут
            best_route = self.select_best_route(group)
            
            if best_route:
                route_html, metadata = best_route
                has_equal_duplicates = len(group) > 1 and any(
                    self.check_rashod_equal_html(r[0]) for r in group
                )
                rashod_equal = self.check_rashod_equal_html(route_html)
                
                # Парсим данные маршрута
                try:
                    route_data = self.parse_html_route(
                        route_html, 
                        metadata,
                        has_equal_duplicates,
                        rashod_equal
                    )
                    
                    # Добавляем количество дубликатов
                    if key in duplicate_counts:
                        if route_data:
                            route_data[0]['Количество дубликатов маршрута'] = duplicate_counts[key]
                    
                    all_rows.extend(route_data)
                    stats['routes_processed'] += 1
                    stats['output_rows'] += len(route_data)
                    
                except Exception as e:
                    logger.error(f"Ошибка обработки маршрута {key}: {e}", exc_info=True)
                    stats['routes_skipped'] += 1
        
        # Создаем DataFrame
        if all_rows:
            df = pd.DataFrame(all_rows)
            logger.info(f"Создан DataFrame с {len(df)} строками")
            return df, stats
        else:
            logger.warning("Нет данных для создания DataFrame")
            return pd.DataFrame(), stats
    
    def _filter_and_select_best_routes(self, all_routes_data: List[Dict]) -> pd.DataFrame:
        """Фильтрует и выбирает лучшие маршруты на основе идентификатора"""
        logger.info("Фильтруем и выбираем лучшие маршруты")
        
        if not all_routes_data:
            return pd.DataFrame()
        
        # Просто создаем DataFrame из всех данных, так как фильтрация уже выполнена
        df = pd.DataFrame(all_routes_data)
        
        logger.info(f"Создан итоговый DataFrame с {len(df)} записями")
        return df
    
    def get_processing_stats(self) -> Dict:
        """Возвращает статистику обработки"""
        return self.processing_stats.copy()
    
    def get_sections_list(self) -> List[str]:
        """Возвращает список участков из обработанных данных"""
        if self.routes_df is None or self.routes_df.empty:
            return []
        
        sections = self.routes_df['Наименование участка'].dropna().unique().tolist()
        logger.debug(f"Найдено участков: {len(sections)}")
        return sorted(sections)
    
    def get_section_data(self, section_name: str) -> pd.DataFrame:
        """Возвращает данные для конкретного участка"""
        if self.routes_df is None or self.routes_df.empty:
            return pd.DataFrame()
        
        section_data = self.routes_df[
            self.routes_df['Наименование участка'] == section_name
        ].copy()
        
        logger.debug(f"Данные для участка '{section_name}': {len(section_data)} записей")
        return section_data
    
    def get_norms_for_section(self, section_name: str) -> List[str]:
        """Возвращает список номеров норм для участка"""
        section_data = self.get_section_data(section_name)
        if section_data.empty:
            return []
        
        norms = section_data['Номер нормы'].dropna().unique().tolist()
        logger.debug(f"Нормы для участка '{section_name}': {norms}")
        return sorted([str(norm) for norm in norms])
    
    def export_to_excel(self, df: pd.DataFrame, output_file: str) -> bool:
        """Экспортирует данные в Excel с форматированием как в route_processor.py"""
        if df.empty:
            logger.warning("DataFrame пуст, нечего экспортировать")
            return False
        
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Border, Side, PatternFill
            
            # Определяем порядок колонок (обновленный с новыми полями)
            columns = [
                'Номер маршрута', 'Дата маршрута', 'Дата поездки', 'Табельный машиниста',
                'Серия локомотива', 'Номер локомотива',
                'НЕТТО', 'БРУТТО', 'ОСИ', 'Наименование участка', 'Номер нормы', 'Дв. тяга',
                'Ткм брутто', 'Км', 'Пр.', 'Расход фактический', 
                'Расход по норме', 'Уд. норма, норма на 1 час ман. раб.', 'Нажатие на ось',
                'Норма на работу', 'Факт уд', 'Факт на работу', 'Норма на одиночное',
                'Простой с бригадой, мин., всего', 'Простой с бригадой, мин., норма',
                'Маневры, мин., всего', 'Маневры, мин., норма',
                'Трогание с места, случ., всего', 'Трогание с места, случ., норма',
                'Нагон опозданий, мин., всего', 'Нагон опозданий, мин., норма',
                'Ограничения скорости, случ., всего', 'Ограничения скорости, случ., норма',
                'На пересылаемые л-вы, всего', 'На пересылаемые л-вы, норма',
                'Количество дубликатов маршрута', 'Н=Ф',
                'USE_RED_COLOR', 'USE_RED_RASHOD'  # Служебные колонки
            ]
            
            # Переупорядочиваем колонки
            existing_columns = [col for col in columns if col in df.columns]
            df_for_excel = df[existing_columns].copy()
            
            # Преобразуем дату
            if 'Дата маршрута' in df_for_excel.columns:
                df_for_excel['Дата маршрута'] = pd.to_datetime(
                    df_for_excel['Дата маршрута'], 
                    format='%d.%m.%Y', 
                    errors='coerce'
                )
            
            # Логируем количество строк с красным цветом
            red_color_count = df_for_excel['USE_RED_COLOR'].sum() if 'USE_RED_COLOR' in df_for_excel.columns else 0
            red_rashod_count = df_for_excel['USE_RED_RASHOD'].sum() if 'USE_RED_RASHOD' in df_for_excel.columns else 0
            
            logger.info(f"Строк с красным цветом для НЕТТО/БРУТТО/ОСИ: {red_color_count}")
            logger.info(f"Строк с красным цветом для расходов: {red_rashod_count}")
            
            # Создаем временный DataFrame без служебных колонок для отображения
            display_columns = [col for col in existing_columns if col not in ['USE_RED_COLOR', 'USE_RED_RASHOD']]
            df_display = df_for_excel[display_columns].copy()
            
            # Сохраняем в Excel
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_display.to_excel(writer, index=False, sheet_name='Маршруты')
                ws = writer.sheets['Маршруты']
                
                # Применяем форматирование (передаем полный DataFrame с флагами)
                self._apply_excel_formatting(ws, df_for_excel)
            
            logger.info(f"Данные экспортированы в {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в Excel: {e}")
            return False
    
    def _apply_excel_formatting(self, ws, df):
        """Применяет форматирование к листу Excel включая красный цвет для приближенных значений"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Border, Side, PatternFill
        
        # Настройки границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thick_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thick')
        )
        
        # Настройки цветов
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        red_font = Font(color='FF0000', bold=True)
        
        # Найдем индексы нужных колонок
        col_indices = {}
        for idx, cell in enumerate(ws[1], 1):  # Первая строка с заголовками
            col_name = cell.value
            if col_name:
                col_indices[col_name] = idx
        
        # Определяем индексы колонок для красного цвета
        netto_col = col_indices.get('НЕТТО')
        brutto_col = col_indices.get('БРУТТО')
        osi_col = col_indices.get('ОСИ')
        rashod_fact_col = col_indices.get('Расход фактический')
        rashod_norm_col = col_indices.get('Расход по норме')
        
        # Применяем форматирование к строкам данных
        for row_idx in range(2, ws.max_row + 1):
            # Определяем границы маршрутов
            route_num = ws.cell(row=row_idx, column=1).value
            next_route = ws.cell(row=row_idx + 1, column=1).value if row_idx < ws.max_row else None
            
            # Получаем флаги из DataFrame
            df_row_idx = row_idx - 2  # Индекс в DataFrame (без заголовка)
            if df_row_idx < len(df):
                use_red_color = df.iloc[df_row_idx].get('USE_RED_COLOR', False)
                use_red_rashod = df.iloc[df_row_idx].get('USE_RED_RASHOD', False)
            else:
                use_red_color = False
                use_red_rashod = False
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Применяем границу
                if route_num != next_route:
                    cell.border = thick_border
                else:
                    cell.border = thin_border
                
                # Применяем красный цвет для НЕТТО, БРУТТО, ОСИ если use_red_color = True
                if use_red_color and col_idx in [netto_col, brutto_col, osi_col]:
                    if col_idx is not None:
                        cell.fill = red_fill
                        cell.font = red_font
                
                # Применяем красный цвет для расходов если use_red_rashod = True
                if use_red_rashod and col_idx in [rashod_fact_col, rashod_norm_col]:
                    if col_idx is not None:
                        cell.fill = red_fill
                        cell.font = red_font
                
                # Форматирование чисел
                if col_idx > 4 and cell.value is not None:  # Пропускаем первые текстовые колонки
                    if isinstance(cell.value, (int, float)) and cell.value != "-":
                        if isinstance(cell.value, int) or cell.value == int(cell.value):
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '#,##0.000'
        
        # Форматирование дат
        date_col = col_indices.get('Дата маршрута')
        if date_col:
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=date_col)
                if date_cell.value:
                    date_cell.number_format = 'DD.MM.YYYY'
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def check_yu6_filter(self, route_html: str) -> bool:
        """Проверяет наличие Ю6 с значениями '1 2 ,0' или '1 3 ,0' для игнорирования маршрута"""
        soup = BeautifulSoup(route_html, 'html.parser')
        
        all_fonts = soup.find_all('font')
        
        for font in all_fonts:
            font_text = font.get_text().strip()
            if font_text.startswith('Ю6'):
                # Проверяем все элементы после Ю6
                current = font
                data_elements = []
                
                while current:
                    current = current.find_next_sibling()
                    if current and current.name == 'font':
                        text = current.get_text().strip().replace('\xa0', ' ').replace('&nbsp;', ' ')
                        if text:
                            data_elements.append(text)
                    elif current and current.name == 'br':
                        break
                    elif not current:
                        break
                
                logger.debug(f"Ю6 данные найдены: {data_elements[:10]}...")
                
                # Ищем паттерны '1 2 ,0' или '1 3 ,0' в данных
                data_str = ' '.join(data_elements)
                if re.search(r'1\s+2\s+,0', data_str) or re.search(r'1\s+3\s+,0', data_str):
                    logger.info("Найден маршрут с Ю6 содержащим '1 2 ,0' или '1 3 ,0' - игнорируется")
                    return True
        
        return False